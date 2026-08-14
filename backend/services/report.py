"""Build comprehensive Excel and PDF financial reports that mirror the app's
analyses (KPIs, net worth, cash flow, spending, budget, investments), gated by a
date range. Reuses the analytics/investments route handlers as plain functions."""
import io
from datetime import date

from sqlalchemy import select, and_
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from schemas import cents_to_display


# ── data gathering ────────────────────────────────────────────────────────────

async def gather_report_data(db: AsyncSession, pid: int, date_from=None, date_to=None,
                             include_investments: bool = True) -> dict:
    """Collect every section the report can show. Analytics handlers are called
    directly with explicit db/pid + date filters (see routers/analytics.py)."""
    from routers.analytics import (
        summary as _summary, by_category as _by_category, cash_flow as _cash_flow,
        net_worth_history as _networth, spending_trends as _trends,
        recurring as _recurring, budget_full as _budget_full,
    )
    from models import Account, Category, Profile

    summ = await _summary(date_from=date_from, date_to=date_to, db=db, pid=pid)
    base_ccy = summ.base_currency
    cat_rows = (await db.execute(select(Category.id, Category.name).where(Category.profile_id == pid))).all()
    prof = await db.get(Profile, pid)
    year = (date_to or date.today()).year

    data = {
        "date_from": date_from, "date_to": date_to, "base_ccy": base_ccy,
        "profile_name": prof.name if prof else None,
        "profile_color": prof.color if prof else None,
        "summary": summ,
        "cash_flow": await _cash_flow(date_from=date_from, date_to=date_to, db=db, pid=pid),
        "by_category": await _by_category(date_from=date_from, date_to=date_to, db=db, pid=pid),
        "trends": await _trends(date_from=date_from, date_to=date_to, db=db, pid=pid),
        "recurring": await _recurring(db=db, pid=pid),
        "budget": await _budget_full(year=year, account_id=None, account_ids=None, db=db, pid=pid),
        "budget_year": year,
        "cat_map": {cid: name for cid, name in cat_rows},
        "top_expenses": await _top_expenses(db, pid, date_from, date_to, base_ccy),
        "investments": None,
        "dividends": None,
    }

    # Net-worth history can hit yfinance for holdings accounts — guard it.
    try:
        data["networth"] = await _networth(date_from=date_from, date_to=date_to, db=db, pid=pid)
    except Exception:
        data["networth"] = None
    accs = (await db.execute(
        select(Account).where(Account.profile_id == pid, Account.is_active == True)  # noqa: E712
    )).scalars().all()
    data["acct_type"] = {a.name: (a.account_type.value if a.account_type else "") for a in accs}

    if include_investments:
        try:
            from routers.investments import investment_accounts as _inv, dividend_calendar as _divcal
            data["investments"] = await _inv(db=db, pid=pid)
            data["dividends"] = await _divcal(months=12, db=db, pid=pid)
        except Exception:
            pass
    return data


async def _top_expenses(db, pid, date_from, date_to, base_ccy, limit=10):
    """Biggest expense transactions in the period, converted to base currency."""
    from models import Transaction
    from services.fx import convert_cents
    filters = [Transaction.profile_id == pid, Transaction.is_debit == True,  # noqa: E712
               Transaction.is_internal_transfer == False]  # noqa: E712
    if date_from:
        filters.append(Transaction.date >= date_from)
    if date_to:
        filters.append(Transaction.date <= date_to)
    stmt = (select(Transaction)
            .options(selectinload(Transaction.account), selectinload(Transaction.category))
            .where(and_(*filters)).order_by(Transaction.amount_cents.desc()).limit(40))
    rows = (await db.execute(stmt)).scalars().all()
    conv = []
    for t in rows:
        c = await convert_cents(db, t.amount_cents, t.currency or "EUR", base_ccy, t.date)
        conv.append((c, t))
    conv.sort(key=lambda x: x[0], reverse=True)
    return [{
        "date": str(t.date), "description": t.description,
        "category": t.category.name if t.category else "—",
        "account": t.account.name if t.account else "—",
        "amount_cents": c,
    } for c, t in conv[:limit]]


# ── shared helpers ────────────────────────────────────────────────────────────

def _eur(cents: int) -> float:
    return round((cents or 0) / 100.0, 2)


def _cell_value(c) -> int:
    planned_active = c.planned_cents != 0 and not c.planned_matched and c.actual_cents == 0
    return c.actual_cents + c.expected_cents + (c.planned_cents if planned_active else 0)


def _total_value(c) -> int:
    return c.actual_cents + c.expected_cents


def _period_label(data) -> str:
    df, dt = data.get("date_from"), data.get("date_to")
    if df and dt:
        return f"{df.strftime('%d/%m/%Y')} — {dt.strftime('%d/%m/%Y')}"
    if dt:
        return f"jusqu'au {dt.strftime('%d/%m/%Y')}"
    return "toutes périodes"


def _summary_rows(summ):
    return [
        ("Patrimoine net", summ.net_worth_cents),
        ("Patrimoine hors emprunts", summ.net_worth_excl_loans_cents),
        ("Emprunts (restant dû)", summ.total_loans_cents),
        ("Revenus (période)", summ.total_income_cents),
        ("Dépenses (période)", summ.total_expenses_cents),
        ("Flux net (période)", summ.net_cash_flow_cents),
    ]


def _patrimoine_by_type(data):
    """{type_label: cents} of current wealth by account type (assets only),
    from the last net-worth entry. Returns None when net worth is unavailable."""
    nw = data.get("networth")
    if not nw:
        return None
    last = nw[-1]
    acct_type = data["acct_type"]
    by = {}
    for k, v in last.items():
        if k in ("month", "total") or (isinstance(k, str) and k.endswith("_native")):
            continue
        t = acct_type.get(k)
        if not t or t == "emprunt":
            continue
        by[t] = by.get(t, 0) + (v or 0)
    by = {t.capitalize(): v for t, v in by.items() if v > 0}
    return by or None


def _rollup_by_category(bc, cat_map):
    """Fold subcategory spending into the parent (single level). No-op until
    CategoryBreakdown carries parent_id. Returns [{name, total_cents, count}] desc."""
    groups, order = {}, []
    for c in bc:
        pid = getattr(c, "parent_id", None)
        top_id = pid if pid is not None else c.category_id
        key = top_id if top_id is not None else "__uncat__"
        if key not in groups:
            name = cat_map.get(top_id) if top_id is not None else c.category_name
            groups[key] = {"name": name or c.category_name, "total_cents": 0, "count": 0}
            order.append(key)
        groups[key]["total_cents"] += c.total_cents
        groups[key]["count"] += c.count
    result = [groups[k] for k in order]
    result.sort(key=lambda g: g["total_cents"], reverse=True)
    return result


def _top_categories_trends(data, n=6):
    """Top-n categories by total spend as series for the trend chart."""
    trends = data.get("trends") or []
    ranked = sorted(trends, key=lambda t: sum(p["amount_cents"] for p in t["series"]), reverse=True)[:n]
    out = []
    for t in ranked:
        out.append({
            "name": t["category_name"],
            "color": t.get("category_color"),
            "months": [p["month"] for p in t["series"]],
            "values_cents": [p["amount_cents"] for p in t["series"]],
        })
    return out


# ── Excel ─────────────────────────────────────────────────────────────────────

def build_xlsx(data, base_ccy: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    bold = Font(bold=True)
    title = Font(bold=True, size=14)
    wb = Workbook()

    # Résumé
    ws = wb.active
    ws.title = "Résumé"
    ws["A1"] = "Résumé financier"
    ws["A1"].font = title
    ws["A2"] = f"Période : {_period_label(data)}"
    ws["A3"] = f"Devise de base : {base_ccy}"
    r = 5
    for label, cents in _summary_rows(data["summary"]):
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=_eur(cents))
        r += 1
    summ = data["summary"]
    if summ.net_worth_by_currency:
        r += 1
        ws.cell(row=r, column=1, value="Patrimoine par devise").font = bold
        r += 1
        for col, head in enumerate(("Devise", "Montant natif", f"Converti ({base_ccy})"), start=1):
            ws.cell(row=r, column=col, value=head).font = bold
        r += 1
        for c in summ.net_worth_by_currency:
            ws.cell(row=r, column=1, value=c.currency)
            ws.cell(row=r, column=2, value=_eur(c.native_cents))
            ws.cell(row=r, column=3, value=_eur(c.converted_cents))
            r += 1
    for col, width in (("A", 28), ("B", 16), ("C", 16)):
        ws.column_dimensions[col].width = width

    # Flux de trésorerie
    wsf = wb.create_sheet("Flux de trésorerie")
    wsf.append(["Mois", "Revenus", "Dépenses", "Flux net"])
    for c in wsf[1]:
        c.font = bold
    for m in data["cash_flow"]:
        wsf.append([m.month, _eur(m.income_cents), _eur(m.expenses_cents), _eur(m.net_cents)])
    wsf.column_dimensions["A"].width = 12

    # Dépenses par catégorie
    wsc = wb.create_sheet("Dépenses par catégorie")
    wsc.append(["Catégorie", "Nb", f"Total ({base_ccy})", "%"])
    for c in wsc[1]:
        c.font = bold
    for cb in data["by_category"]:
        wsc.append([cb.category_name, cb.count, _eur(cb.total_cents), round(cb.percentage, 1)])
    wsc.column_dimensions["A"].width = 26

    # Top dépenses
    wst = wb.create_sheet("Top dépenses")
    wst.append(["Date", "Description", "Catégorie", "Compte", f"Montant ({base_ccy})"])
    for c in wst[1]:
        c.font = bold
    for t in data["top_expenses"]:
        wst.append([t["date"], t["description"], t["category"], t["account"], _eur(t["amount_cents"])])
    for col, w in (("A", 12), ("B", 44), ("C", 20), ("D", 16), ("E", 16)):
        wst.column_dimensions[col].width = w

    # Récurrents
    wsr = wb.create_sheet("Récurrents")
    wsr.append(["Description", "Catégorie", "Occurrences", "Montant moyen", "Dernière date"])
    for c in wsr[1]:
        c.font = bold
    cat_map = data["cat_map"]
    for rec in data["recurring"]:
        wsr.append([rec.description, cat_map.get(rec.category_id, "—"), rec.occurrences,
                    _eur(rec.avg_amount_cents), str(rec.last_date)])
    for col, w in (("A", 44), ("B", 20), ("C", 12), ("D", 16), ("E", 14)):
        wsr.column_dimensions[col].width = w

    # Budget
    _xlsx_budget_sheet(wb, data["budget"], bold)

    # Investments
    if data.get("investments"):
        _xlsx_investments(wb, data, bold)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_budget_sheet(wb, budget, bold):
    ws = wb.create_sheet("Budget")
    ws.append(["Catégorie", *budget.months, "Total"])
    for cell in ws[1]:
        cell.font = bold
    for section in budget.sections:
        ws.append([section.section_label])
        ws.cell(row=ws.max_row, column=1).font = bold
        for row in section.rows:
            total = sum(_cell_value(c) for c in row.cells)
            ws.append([row.category_name, *[_eur(_cell_value(c)) for c in row.cells], _eur(total)])
        tot = section.section_totals
        total = sum(_total_value(c) for c in tot.cells)
        ws.append([tot.category_name, *[_eur(_total_value(c)) for c in tot.cells], _eur(total)])
        for cell in ws[ws.max_row]:
            cell.font = bold
    ws.column_dimensions["A"].width = 26


def _xlsx_investments(wb, data, bold):
    wsp = wb.create_sheet("Positions")
    wsp.append(["Compte", "Ticker", "Nom", "Type", "Qté", "Prix", "Valeur", "Coût", "+/-", "%"])
    for c in wsp[1]:
        c.font = bold
    for acc in data["investments"]:
        for h in (acc.get("holdings") or []):
            wsp.append([
                acc["name"], h.get("ticker"), h.get("name"), h.get("asset_type"),
                round((h.get("quantity") or 0), 4),
                _eur(h.get("current_price_cents") or 0),
                _eur(h.get("value_in_account_ccy_cents") or h.get("current_value_cents") or 0),
                _eur(h.get("cost_basis_cents") or 0),
                _eur(h.get("gain_cents") or 0),
                round((h.get("gain_pct") or 0), 1),
            ])
    for col, w in (("A", 16), ("B", 12), ("C", 30), ("D", 12)):
        wsp.column_dimensions[col].width = w

    div = data.get("dividends") or {}
    wsd = wb.create_sheet("Dividendes")
    wsd.append(["Secteur", "Revenu annuel estimé"])
    for c in wsd[1]:
        c.font = bold
    for s in div.get("by_sector", []):
        wsd.append([s.get("sector") or "—", _eur(s.get("est_annual_cents") or 0)])
    wsd.append([])
    wsd.append(["Positions versant un dividende"])
    wsd.cell(row=wsd.max_row, column=1).font = bold
    wsd.append(["Ticker", "Nom", "Qté", "Rendement %", "Revenu est./an", "Fréq."])
    for c in wsd[wsd.max_row]:
        c.font = bold
    for acc in data["investments"]:
        for h in (acc.get("holdings") or []):
            if (h.get("est_annual_income_cents") or 0) > 0 or (h.get("dividend_yield") or 0) > 0:
                wsd.append([
                    h.get("ticker"), h.get("name"), round((h.get("quantity") or 0), 4),
                    round((h.get("dividend_yield") or 0), 2),
                    _eur(h.get("est_annual_income_cents") or 0), h.get("frequency") or "—",
                ])
    wsd.column_dimensions["A"].width = 12
    wsd.column_dimensions["B"].width = 30


# ── PDF ───────────────────────────────────────────────────────────────────────

def build_pdf(data, base_ccy: str, opts=None) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, KeepTogether
    from services import report_charts as rc

    REG, BOLD = rc.register_pdf_fonts()
    base = getSampleStyleSheet()
    st_title = ParagraphStyle("t", parent=base["Title"], fontName=BOLD)
    st_h2 = ParagraphStyle("h2", parent=base["Heading2"], fontName=BOLD, spaceBefore=10)
    st_n = ParagraphStyle("n", parent=base["Normal"], fontName=REG)
    st_muted = ParagraphStyle("m", parent=st_n, textColor=colors.HexColor("#64748b"))

    PAGE_W = 27.7  # landscape A4 usable cm

    def png_image(png: bytes, width_cm: float) -> Image:
        im = Image(io.BytesIO(png))
        im.drawWidth = width_cm * cm
        im.drawHeight = width_cm * cm * im.imageHeight / im.imageWidth
        return im

    def styled_table(rows, col_widths=None, header=True, font_size=8, align_right_from=1):
        t = Table(rows, colWidths=col_widths, repeatRows=1 if header else 0)
        cmds = [
            ("FONTNAME", (0, 0), (-1, -1), REG),
            ("FONTSIZE", (0, 0), (-1, -1), font_size),
            ("ALIGN", (align_right_from, 0), (-1, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
        ]
        if header:
            cmds += [
                ("FONTNAME", (0, 0), (-1, 0), BOLD),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
                ("ALIGN", (0, 0), (-1, 0), "LEFT"),
            ]
        t.setStyle(TableStyle(cmds))
        return t

    def money(cents):
        return cents_to_display(cents, base_ccy)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Rapport financier",
                            leftMargin=1 * cm, rightMargin=1 * cm, topMargin=1.7 * cm, bottomMargin=1.4 * cm)
    summ = data["summary"]

    # Repeating header (profile · title) + footer (page · date · period) on every page.
    profile_name = data.get("profile_name") or "Rapport financier"
    period = _period_label(data)
    gen_date = date.today().strftime("%d/%m/%Y")

    def _hf(canvas, docu):
        canvas.saveState()
        w, h = landscape(A4)
        canvas.setFont(REG, 8)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(1 * cm, h - 1.0 * cm, profile_name)
        canvas.drawRightString(w - 1 * cm, h - 1.0 * cm, "Rapport financier")
        canvas.setStrokeColor(colors.HexColor("#e5e7eb"))
        canvas.line(1 * cm, h - 1.15 * cm, w - 1 * cm, h - 1.15 * cm)
        canvas.drawCentredString(w / 2, 0.7 * cm, f"Page {docu.page} · généré le {gen_date} · {period}")
        canvas.restoreState()

    def sect(*flowables):
        """Keep a section title with its chart/table so a heading never orphans."""
        return KeepTogether(list(flowables))

    e = [
        Paragraph("Rapport financier", st_title),
        Paragraph(f"{profile_name} · Période : {period} · Devise de base : {base_ccy}", st_muted),
        Spacer(1, 0.4 * cm),
    ]

    # 1. KPIs
    kpi_flow = [Paragraph("Vue d'ensemble", st_h2),
                styled_table([[label, money(cents)] for label, cents in _summary_rows(summ)],
                             col_widths=[7 * cm, 5 * cm], header=False)]
    if summ.net_worth_by_currency:
        rows = [["Devise", "Montant natif", f"Converti ({base_ccy})"]]
        for c in summ.net_worth_by_currency:
            rows.append([c.currency, cents_to_display(c.native_cents, c.currency), money(c.converted_cents)])
        kpi_flow += [Spacer(1, 0.3 * cm), styled_table(rows, col_widths=[4 * cm, 6 * cm, 6 * cm])]
    e.append(sect(*kpi_flow))

    # 2. Net worth
    if data.get("networth"):
        nw = data["networth"]
        months = [row["month"] for row in nw]
        totals = [int(row.get("total") or 0) for row in nw]
        e.append(sect(Paragraph("Évolution du patrimoine", st_h2),
                      png_image(rc.networth_area(months, totals, base_ccy), PAGE_W * 0.62)))
        by_type = _patrimoine_by_type(data)
        if by_type:
            e.append(sect(Paragraph("Répartition du patrimoine par type de compte", st_h2),
                          png_image(rc.donut(list(by_type.keys()), list(by_type.values()), base_ccy), 13)))

    # 3. Cash flow
    cf = data.get("cash_flow") or []
    if cf:
        months = [m.month for m in cf]
        e.append(sect(Paragraph("Revenus & dépenses", st_h2), png_image(rc.cashflow_bars(
            months, [m.income_cents for m in cf], [m.expenses_cents for m in cf],
            [m.net_cents for m in cf], base_ccy), PAGE_W * 0.66)))

    # 4. Spending — rolled up by parent category (children folded into their parent)
    bc = _rollup_by_category(data.get("by_category") or [], data["cat_map"])
    if bc:
        top = bc[:6]
        others = sum(c["total_cents"] for c in bc[6:])
        labels = [c["name"] for c in top] + (["Autres"] if others > 0 else [])
        values = [c["total_cents"] for c in top] + ([others] if others > 0 else [])
        rows = [["Catégorie", "Nb", f"Total ({base_ccy})", "%"]]
        grand = sum(c["total_cents"] for c in bc) or 1
        for c in bc:
            rows.append([c["name"], str(c["count"]), money(c["total_cents"]), f"{c['total_cents'] / grand * 100:.1f}%"])
        e.append(sect(Paragraph("Dépenses par catégorie", st_h2),
                      png_image(rc.donut(labels, values, base_ccy), 13),
                      styled_table(rows, col_widths=[10 * cm, 3 * cm, 6 * cm, 3 * cm])))

    # 5. Top expenses
    te = data.get("top_expenses") or []
    if te:
        rows = [["Date", "Description", "Catégorie", "Compte", f"Montant ({base_ccy})"]]
        for t in te:
            rows.append([t["date"], t["description"][:60], t["category"], t["account"], money(t["amount_cents"])])
        e.append(sect(Paragraph("Top 10 des dépenses", st_h2),
                      styled_table(rows, col_widths=[2.6 * cm, 13 * cm, 4.5 * cm, 3.5 * cm, 3.6 * cm])))

    # 6. Category trends
    tr = _top_categories_trends(data)
    if tr:
        e.append(sect(Paragraph("Tendances par catégorie", st_h2),
                      png_image(rc.category_trends(tr, base_ccy), PAGE_W * 0.66)))

    # 7. Recurring
    rec = data.get("recurring") or []
    if rec:
        cat_map = data["cat_map"]
        rows = [["Description", "Catégorie", "Occ.", "Montant moyen", "Dernière"]]
        for r in rec[:20]:
            rows.append([r.description[:55], cat_map.get(r.category_id, "—"), str(r.occurrences),
                         cents_to_display(r.avg_amount_cents, base_ccy), str(r.last_date)])
        e.append(sect(Paragraph("Transactions récurrentes", st_h2),
                      styled_table(rows, col_widths=[11 * cm, 5 * cm, 2 * cm, 4 * cm, 3 * cm])))

    # 8. Budget
    e.append(sect(Paragraph(f"Budget {data['budget_year']}", st_h2),
                  _pdf_budget_table(data["budget"], REG, BOLD)))

    # 9. Investments
    if data.get("investments"):
        _pdf_investments(e, data, base_ccy, png_image, styled_table, st_h2, rc, REG, BOLD)

    doc.build(e, onFirstPage=_hf, onLaterPages=_hf)
    return buf.getvalue()


def _pdf_budget_table(budget, REG, BOLD):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Table, TableStyle

    def fmt(cents):
        return f"{round((cents or 0) / 100):,}".replace(",", " ")

    def short(m):
        return f"{m[5:7]}/{m[2:4]}"

    header = ["Catégorie", *[short(m) for m in budget.months], "Total"]
    rows = [header]
    cmds = [
        ("FONTNAME", (0, 0), (-1, -1), REG),
        ("FONTSIZE", (0, 0), (-1, -1), 6),
        ("FONTNAME", (0, 0), (-1, 0), BOLD),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
    ]
    ridx = 1
    for section in budget.sections:
        rows.append([section.section_label, *[""] * (len(budget.months) + 1)])
        cmds += [("SPAN", (0, ridx), (-1, ridx)), ("FONTNAME", (0, ridx), (-1, ridx), BOLD),
                 ("BACKGROUND", (0, ridx), (-1, ridx), colors.HexColor("#f3f4f6"))]
        ridx += 1
        for row in section.rows:
            total = sum(_cell_value(c) for c in row.cells)
            rows.append([row.category_name, *[fmt(_cell_value(c)) for c in row.cells], fmt(total)])
            ridx += 1
        tot = section.section_totals
        total = sum(_total_value(c) for c in tot.cells)
        rows.append([tot.category_name, *[fmt(_total_value(c)) for c in tot.cells], fmt(total)])
        cmds.append(("FONTNAME", (0, ridx), (-1, ridx), BOLD))
        ridx += 1

    n = len(budget.months)
    month_w = (27.7 - 3.2 - 1.8) / max(1, n)
    col_widths = [3.2 * cm, *[month_w * cm] * n, 1.8 * cm]
    t = Table(rows, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(cmds))
    return t


def _pdf_investments(e, data, base_ccy, png_image, styled_table, st_h2, rc, REG, BOLD):
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    accs = data["investments"]
    st_n = ParagraphStyle("invn", parent=getSampleStyleSheet()["Normal"], fontName=REG)

    def money(cents):
        return cents_to_display(cents, base_ccy)

    e.append(Paragraph("Investissements", st_h2))
    total_val = sum(a.get("current_value_cents") or 0 for a in accs)
    e.append(Paragraph(f"Valeur totale : {money(total_val)}", st_n))

    # Global allocation by asset type
    alloc = {}
    for a in accs:
        for t, v in (a.get("allocation_by_type") or {}).items():
            alloc[t] = alloc.get(t, 0) + (v or 0)
    alloc = {k: v for k, v in alloc.items() if v > 0}
    if alloc:
        e.append(png_image(rc.donut(list(alloc.keys()), list(alloc.values()), base_ccy), 13))

    # Per-account holdings
    for a in accs:
        holdings = a.get("holdings") or []
        if not holdings:
            continue
        e.append(Paragraph(f"{a['name']} — {money(a.get('holdings_value_cents') or 0)} "
                           f"({a.get('currency', '')})", st_h2))
        rows = [["Ticker", "Nom", "Type", "Qté", "Valeur", "Coût", "+/-", "%"]]
        for h in holdings:
            rows.append([
                h.get("ticker") or "", (h.get("name") or "")[:34], h.get("asset_type") or "",
                f"{h.get('quantity') or 0:g}",
                cents_to_display(h.get("value_in_account_ccy_cents") or h.get("current_value_cents") or 0, a.get("currency", base_ccy)),
                cents_to_display(h.get("cost_basis_cents") or 0, a.get("currency", base_ccy)),
                cents_to_display(h.get("gain_cents") or 0, a.get("currency", base_ccy)),
                f"{h.get('gain_pct') or 0:.1f}%",
            ])
        e.append(styled_table(rows, col_widths=[2.5 * cm, 8 * cm, 2.5 * cm, 2.5 * cm, 3.5 * cm, 3.5 * cm, 3.2 * cm, 2 * cm]))

    # Dividends
    div = data.get("dividends") or {}
    monthly = div.get("monthly") or []
    if monthly:
        # Build per-sector monthly series for a stacked bar.
        months = [m["month"] for m in monthly]
        sector_totals = {}
        for m in monthly:
            for it in m.get("items", []):
                sec = it.get("sector") or "Autre"
                sector_totals.setdefault(sec, {})[m["month"]] = sector_totals.get(sec, {}).get(m["month"], 0) + (it.get("amount_cents") or 0)
        sectors = [{"name": sec, "values_cents": [vals.get(mm, 0) for mm in months]}
                   for sec, vals in sector_totals.items()]
        if any(any(v for v in s["values_cents"]) for s in sectors):
            e.append(Paragraph("Projection des dividendes (12 mois)", st_h2))
            e.append(png_image(rc.dividend_stacked(months, sectors, base_ccy), 27.7 * 0.66))

    positions = []
    for a in accs:
        for h in (a.get("holdings") or []):
            if (h.get("est_annual_income_cents") or 0) > 0:
                positions.append((a, h))
    if positions:
        e.append(Paragraph("Positions versant un dividende", st_h2))
        rows = [["Ticker", "Nom", "Rendement", "Revenu est./an", "Fréq."]]
        for a, h in positions:
            rows.append([
                h.get("ticker") or "", (h.get("name") or "")[:34],
                f"{h.get('dividend_yield') or 0:.2f}%",
                cents_to_display(h.get("est_annual_income_cents") or 0, a.get("currency", base_ccy)),
                h.get("frequency") or "—",
            ])
        e.append(styled_table(rows, col_widths=[2.5 * cm, 10 * cm, 3 * cm, 4 * cm, 3 * cm]))
